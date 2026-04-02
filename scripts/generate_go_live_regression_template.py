# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "用例ID",
    "阶段",
    "模块",
    "检查项",
    "优先级",
    "阻塞上线",
    "执行状态",
    "执行人",
    "执行日期",
    "证据链接/截图说明",
    "缺陷单号",
    "结果说明",
    "下一步动作",
]

ROWS = [
    ["GLR-001", "发布准备", "发布中心", "选择目标工作场景，确认场景状态、账期和默认版本正确", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-002", "发布准备", "发布中心", "完成一次正式发布，确认版本号、发布时间、发布人和快照摘要可见", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-003", "发布准备", "发布快照", "校验影响因素、规则、价格方案、变量配置均进入发布快照，且详情抽屉可查看", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-004", "正式核算", "任务中心", "使用 JSON 直传提交单笔正式核算任务，并生成结果", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-005", "正式核算", "任务中心", "使用 JSON 直传提交批量输入，确认任务明细、结果台账和追溯台账正常生成", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-006", "正式核算", "导入批次", "创建导入批次，确认批次台账可查看批次摘要、样例明细和状态", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-007", "正式核算", "导入批次", "从导入批次台账回填批次继续提交正式核算任务，确认任务正确引用 sourceBatchNo", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-008", "分片恢复", "任务详情", "提交一笔足够触发多分片的正式核算任务，确认分片列表、范围、耗时和成功失败数正确", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-009", "分片恢复", "任务详情", "制造至少一条失败明细，确认失败聚合摘要和高频错误原因展示正确", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-010", "分片恢复", "分片监控", "对失败分片执行一次重试，确认分片状态、任务统计和任务明细同步刷新", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-011", "分片恢复", "任务详情", "对失败明细执行一次重试，确认兼容链路可用且不会覆盖其他分片结果", "P0", "否", "待执行", "", "", "", "", "", ""],
    ["GLR-012", "治理联动", "告警中心", "从任务中心进入告警中心，确认自动带入 sceneId、billMonth、taskId 和未关闭状态", "P0", "否", "待执行", "", "", "", "", "", ""],
    ["GLR-013", "治理联动", "告警中心", "任务失败或重试超限时，确认告警台账能看到记录，并可执行确认、关闭", "P0", "否", "待执行", "", "", "", "", "", ""],
    ["GLR-014", "治理联动", "缓存治理", "校验运行快照缓存状态可查询，缓存刷新有权限控制且执行后仍以数据库为准", "P0", "否", "待执行", "", "", "", "", "", ""],
    ["GLR-015", "结果追溯", "结果台账", "按场景、账期、任务过滤结果，确认金额、对象标识和执行版本正确", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-016", "结果追溯", "追溯解释", "打开结果详情和追溯详情，确认规则命中、变量快照和解释链路可查看", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-017", "结果追溯", "结果台账", "校验导入批次任务与 JSON 任务在结果台账中的查询口径一致", "P0", "是", "待执行", "", "", "", "", "", ""],
    ["GLR-018", "权限验证", "菜单权限", "管理员可在左侧菜单看到导入批次、正式核算、结果台账、告警中心", "P0", "是", "待环境验证", "", "", "", "", "", ""],
    ["GLR-019", "权限验证", "角色隔离", "普通角色在未授权 cost:task:list、cost:alarm:list 等权限时，不应看到对应菜单与操作按钮", "P0", "是", "待环境验证", "", "", "", "", "", ""],
    ["GLR-020", "权限验证", "按钮权限", "任务取消、分片重试、告警确认/关闭、缓存刷新均受按钮权限控制", "P0", "否", "待执行", "", "", "", "", "", ""],
    ["GLR-021", "环境验证", "Flyway", "目标环境执行 Flyway 后，确认 V20260402_012/013/014 迁移按顺序成功", "P0", "是", "待环境验证", "", "", "", "", "", ""],
    ["GLR-022", "环境验证", "构建校验", "后端执行 mvn -q -DskipTests package 通过", "P0", "否", "已通过", "", "", "本地已完成", "", "本地构建已通过", "上线前在目标分支再执行一次"],
    ["GLR-023", "环境验证", "构建校验", "前端执行 npm run build:prod 通过", "P0", "否", "已通过", "", "", "本地已完成", "", "本地构建已通过", "上线前在目标分支再执行一次"],
]

STATUS_FILL = {
    "待执行": "FFF2CC",
    "待环境验证": "F4CCCC",
    "已通过": "E2F0D9",
    "已失败": "F4CCCC",
    "已阻塞": "FCE4D6",
}

PRIORITY_FILL = {
    "P0": "CFE2F3",
    "P1": "D9EAD3",
    "P2": "FFF2CC",
    "P3": "F4CCCC",
}


def build_template(output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ExecutionLedger"
    sheet.append(HEADERS)
    for row in ROWS:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        priority = sheet.cell(row_idx, 5).value
        blocking = sheet.cell(row_idx, 6).value
        status = sheet.cell(row_idx, 7).value
        if priority in PRIORITY_FILL:
            sheet.cell(row_idx, 5).fill = PatternFill("solid", fgColor=PRIORITY_FILL[priority])
        sheet.cell(row_idx, 6).fill = PatternFill("solid", fgColor="F4CCCC" if blocking == "是" else "E2F0D9")
        if status in STATUS_FILL:
            sheet.cell(row_idx, 7).fill = PatternFill("solid", fgColor=STATUS_FILL[status])

    widths = {
        1: 12,
        2: 12,
        3: 12,
        4: 34,
        5: 8,
        6: 10,
        7: 12,
        8: 12,
        9: 14,
        10: 28,
        11: 16,
        12: 28,
        13: 24,
    }
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A2"

    table = Table(displayName="ExecutionLedgerTable", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    summary = workbook.create_sheet("Summary")
    summary.append(["维度", "值"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    summary_rows = [
        ["总用例数", len(ROWS)],
        ["阻塞上线用例数", sum(1 for row in ROWS if row[5] == "是")],
        ["待执行", sum(1 for row in ROWS if row[6] == "待执行")],
        ["待环境验证", sum(1 for row in ROWS if row[6] == "待环境验证")],
        ["已通过", sum(1 for row in ROWS if row[6] == "已通过")],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 16

    guide = workbook.create_sheet("Guide")
    guide.append(["字段", "说明"])
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    guide_rows = [
        ["执行状态", "建议使用：待执行、待环境验证、已通过、已失败、已阻塞。"],
        ["证据链接/截图说明", "可填写截图文件名、测试地址、日志路径、数据库截图位置或工单链接。"],
        ["缺陷单号", "若执行失败，请绑定缺陷单或问题单，方便回归闭环。"],
        ["结果说明", "填写实际执行结论，例如‘与预期一致’、‘菜单未显示’、‘分片重试后统计未刷新’。"],
        ["下一步动作", "明确后续责任动作，例如‘开发修复后回归’、‘DBA 执行 Flyway’、‘测试补充证据’。"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 90
    for row_idx in range(2, guide.max_row + 1):
        for col_idx in range(1, 3):
            guide.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
            guide.cell(row_idx, col_idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    workbook.save(output_path)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    output = root / "docs" / "go_live_regression_execution_template.xlsx"
    build_template(output)
    workbook = load_workbook(output, read_only=True, data_only=True)
    print(output)
    print("sheets=", workbook.sheetnames)
    print("rows=", workbook["ExecutionLedger"].max_row - 1)


if __name__ == "__main__":
    main()
